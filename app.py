from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory, jsonify, make_response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from pymongo import MongoClient
from pymongo.errors import ConnectionFailure
import os
from datetime import datetime
import openpyxl
from dotenv import load_dotenv
import math
import re
import tempfile
try:
    from weasyprint import HTML
except ImportError:
    HTML = None
from PIL import Image

# Carrega variáveis de ambiente
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'chave-secreta-padrao')

# Configuração do MongoDB
client = None
db = None
produtos_collection = None

try:
    mongodb_uri = os.getenv('MONGODB_URI')
    if not mongodb_uri:
        raise ValueError("MONGODB_URI não encontrada nas variáveis de ambiente")
    
    client = MongoClient(mongodb_uri)
    # Testa a conexão
    client.admin.command('ping')
    db = client['catalogo_produtos']
    produtos_collection = db['produtos']
    print("Conectado ao MongoDB com sucesso!")
except Exception as e:
    print(f"Erro ao conectar ao MongoDB: {str(e)}")
    # Em um ambiente de produção, você provavelmente não mostraria o erro diretamente
    # e lidaria com ele de forma mais robusta.
    # raise # Manter para debug em desenvolvimento, mas remover em produção

# Configuração do Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Configuração de upload
UPLOAD_FOLDER = 'static/uploads/images'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Classe de usuário para Flask-Login
class User(UserMixin):
    def __init__(self, id, role):
        self.id = id
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    if db is None:
        return None
    
    user_data = db.users.find_one({'username': user_id})
    if user_data:
        return User(user_data['username'], user_data['role'])
    return None

# Funções auxiliares
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Rotas
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'dono':
            return redirect(url_for('catalogo'))
        else:
            return redirect(url_for('galeria'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if db is None:
            flash('Erro: Conexão com o banco de dados não estabelecida.')
            return render_template('login.html')
        
        # Buscar usuário no banco de dados
        user_data = db.users.find_one({'username': username})
        
        if user_data and check_password_hash(user_data['password'], password):
            user = User(user_data['username'], user_data['role'])
            login_user(user)
            return redirect(url_for('galeria'))
        
        flash('Credenciais inválidas')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você foi desconectado com sucesso.')
    return redirect(url_for('login'))

@app.route('/galeria')
@login_required
def galeria():
    if produtos_collection is None:
         flash('Erro: Conexão com o MongoDB não estabelecida.')
         return render_template('galeria.html', produtos=[], total_pages=0, current_page=1, search_query='', tabela_preco_selecionada='TABELA_VAREJO_5_000', categorias=[])

    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '')
    tabela_preco_selecionada = request.args.get('tabela_preco', 'TABELA_VAREJO_5_000')
    categoria_selecionada = request.args.get('categoria', '')
    cor_selecionada = request.args.get('cor', '')
    per_page = 24

    # Buscar todas as categorias únicas do banco
    try:
        categorias = produtos_collection.distinct('Desc_Familia')
        categorias = [cat for cat in categorias if cat]
        categorias.sort()
    except Exception as e:
        print(f"Erro ao buscar categorias: {str(e)}")
        categorias = []

    # Buscar todas as cores únicas do banco
    try:
        cores = produtos_collection.distinct('Desc_Cor')
        cores = [cor for cor in cores if cor]
        cores.sort()
    except Exception as e:
        print(f"Erro ao buscar cores: {str(e)}")
        cores = []

    # Pipeline de agregação para agrupar por código base e incluir variantes
    pipeline = []

    # Estágio 1: Filtro inicial (ativo_catalogo, busca e categoria)
    match_query = {}
    if not (current_user.is_authenticated and current_user.role == 'dono'):
        match_query['ativo_catalogo'] = True

    if search_query:
        match_query['descricao_produto'] = {'$regex': re.escape(search_query), '$options': 'i'}
    
    if categoria_selecionada:
        match_query['Desc_Familia'] = categoria_selecionada
        
    pipeline.append({'$match': match_query})

    # Estágio 2: Adicionar codigo_base
    # Remove o último caractere do codigo_produto para criar o código base
    pipeline.append({
        '$addFields': {
            'codigo_base': { '$substrCP': ['$codigo_produto', 0, { '$subtract': [{'$strLenCP': '$codigo_produto'}, 1] }] }
        }
    })
    
    # Estágio 3: Agrupar por codigo_base e coletar variantes
    # Também encontrar o produto 'principal' (BRANCO) para o card
    pipeline.append({
        '$group': {
            '_id': '$codigo_base',
            'variants': { '$push': '$$ROOT' },
            'main_product': { # Tenta encontrar o produto BRANCO (código termina em 1 ou Desc_Cor é BRANCO)
                '$first': { # Usamos $first após ordenar/filtrar implicitamente na lista 'variants'
                    '$filter': {
                        'input': '$variants',
                        'as': 'variant',
                        'cond': { '$or': [
                            { '$eq': [{ '$substrCP': [{'$ifNull': ['$$variant.codigo_produto', '']}, { '$subtract': [{'$strLenCP': {'$ifNull': ['$$variant.codigo_produto', '']}}, 1] }, 1] }, '1'] },
                            { '$eq': [{'$ifNull': ['$$variant.Desc_Cor', '']}, 'BRANCO'] }
                        ]}
                    }
                 }
            }
        }
    })
    
    # Estágio 4: Projetar o resultado para ter a estrutura desejada
    # Usar o main_product encontrado, ou o primeiro da lista de variantes se nenhum BRANCO for encontrado
    # Note: $first dentro de $group pode retornar null se o filtro não encontrar nada. 
    # Precisamos garantir que main_product seja um objeto válido.
    pipeline.append({
        '$project': {
            '_id': 0, # Remove o _id do grupo
            'codigo_base': '$_id', # Renomeia _id do grupo para codigo_base
            'main_product': { 
                '$ifNull': ['$main_product', { '$arrayElemAt': ['$variants', 0] }] } # Usa main_product se existir, senão o primeiro da lista
            ,
            'variants': '$variants' # Mantém a lista completa de variantes
        }
    })
    
    # Precisamos contar o total de produtos agrupados *antes* de aplicar skip/limit para paginação correta
    # Criar um pipeline separado para a contagem
    count_pipeline = pipeline[:-1] # Usa todos os estágios exceto o último ($project)
    count_pipeline.append({'$count': 'total'})

    total_produtos_agrupados = 0
    try:
        count_result = list(produtos_collection.aggregate(count_pipeline))
        if count_result:
            total_produtos_agrupados = count_result[0].get('total', 0)
    except Exception as e:
        print(f"Erro ao contar produtos agrupados: {str(e)}")
        # Continua com 0 se houver erro na contagem

    total_pages = math.ceil(total_produtos_agrupados / per_page)

    # Garante que a página solicitada é válida
    if page < 1:
        page = 1
    elif page > total_pages and total_pages > 0:
        page = total_pages
    elif total_pages == 0: # Se não houver produtos, a página é 1 mas o skip/limit não trará resultados
         page = 1

    skip = (page - 1) * per_page
    
    # Estágios para paginação
    pipeline.append({'$skip': skip})
    pipeline.append({'$limit': per_page})
    
    # Executar o pipeline de agregação
    produtos_agrupados = []
    try:
        produtos_agrupados = list(produtos_collection.aggregate(pipeline))
        # Processar os resultados para ter uma estrutura mais amigável no template
        produtos_para_template = []
        for grupo in produtos_agrupados:
            # Se cor_selecionada, só adiciona o grupo se alguma variante tem essa cor
            if cor_selecionada:
                if not any(variant.get('Desc_Cor') == cor_selecionada for variant in grupo['variants']):
                    continue
            main_prod = grupo['main_product']
            variants_list = []
            for variant in grupo['variants']:
                tabelas_preco = [
                    'TABELA_VAREJO_5_000',
                    'TABELA_VAREJO_25_000',
                    'TABELA_VAREJO_50_000',
                    'TABELA_ATACADO',
                    'TABELA_GOLD_PARCEIRO',
                    'TABELA_ATACADO_AMAZONIA_E_MACAPA',
                    'TABELA_ACRE_E_RONDONIA'
                ]
                precos = {k: variant.get(k, 0) for k in tabelas_preco}
                variant_data = {
                    'codigo_produto': variant.get('codigo_produto'),
                    'Desc_Cor': variant.get('Desc_Cor', 'Não Informado'),
                    'precos': precos
                }
                variants_list.append(variant_data)
            produtos_para_template.append({
                'codigo_base': grupo['codigo_base'],
                'descricao_produto': main_prod.get('descricao_produto'),
                'caminho_imagem': main_prod.get('caminho_imagem'),
                'ativo_catalogo': main_prod.get('ativo_catalogo', False),
                'variants': variants_list
            })
        return render_template('galeria.html', 
                               produtos=produtos_para_template, 
                               total_pages=total_pages, 
                               current_page=page,
                               search_query=search_query,
                               tabela_preco_selecionada=tabela_preco_selecionada,
                               categorias=categorias,
                               categoria_selecionada=categoria_selecionada,
                               cores=cores,
                               cor_selecionada=cor_selecionada)
    except Exception as e:
        print("ERRO GRAVE NA GALERIA:", e)
        import traceback
        traceback.print_exc()
        flash(f'Erro ao carregar produtos na galeria: {str(e)}')
        return render_template('galeria.html', produtos=[], total_pages=0, current_page=1, search_query=search_query, tabela_preco_selecionada=tabela_preco_selecionada, categorias=categorias, categoria_selecionada=categoria_selecionada)

@app.route('/admin')
@login_required
def admin():
    if current_user.role != 'admin':
        flash('Acesso não autorizado')
        return redirect(url_for('galeria'))
    
    if produtos_collection is None:
         flash('Erro: Conexão com o MongoDB não estabelecida.')
         return render_template('admin.html', produtos=[], total_pages=0, current_page=1, search_query='')

    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '')
    per_page = 50

    query = {}
    if search_query:
        # Busca por descrição (case-insensitive)
        # Usar re.escape para garantir que caracteres especiais no termo de busca sejam tratados
        query['descricao_produto'] = {'$regex': re.escape(search_query), '$options': 'i'}

    try:
        total_produtos = produtos_collection.count_documents(query)
        total_pages = math.ceil(total_produtos / per_page)
        
        # Garante que a página solicitada é válida
        if page < 1:
            page = 1
        elif page > total_pages and total_pages > 0:
            page = total_pages
        elif total_pages == 0:
             page = 1

        skip = (page - 1) * per_page
        produtos = list(produtos_collection.find(query).skip(skip).limit(per_page))

        print("DEBUG: Resultado da consulta /admin:", list(produtos_collection.find(query).skip(skip).limit(per_page)))

        return render_template('admin.html', 
                               produtos=produtos, 
                               total_pages=total_pages, 
                               current_page=page,
                               search_query=search_query)
    except Exception as e:
        flash(f'Erro ao carregar produtos no admin: {str(e)}')
        print(f"Erro na rota /admin: {str(e)}")
        return render_template('admin.html', produtos=[], total_pages=0, current_page=1, search_query=search_query)

# Nova rota para buscar produtos para upload de imagem
@app.route('/search_products_for_upload', methods=['GET'])
@login_required
def search_products_for_upload():
    if current_user.role != 'admin':
        return jsonify({'error': 'Acesso não autorizado'}), 403

    if produtos_collection is None:
         return jsonify({'error': 'Erro: Conexão com o MongoDB não estabelecida.'}), 500

    search_term = request.args.get('description', '')

    if not search_term:
        return jsonify([]) # Retorna lista vazia se não houver termo de busca

    try:
        # Busca por descrição (case-insensitive) e limita o número de resultados para não sobrecarregar
        # Usar re.escape para garantir que caracteres especiais no termo de busca sejam tratados
        query = {'descricao_produto': {'$regex': re.escape(search_term), '$options': 'i'}}
        # Buscar apenas produtos que ainda não tem imagem associada (opcional, mas útil)
        # query['caminho_imagem'] = {'$exists': False}

        produtos = list(produtos_collection.find(query).limit(20)) # Limita a 20 resultados
        
        # Formatar resultados para JSON
        results = []
        for produto in produtos:
            results.append({
                'codigo_produto': produto.get('codigo_produto'),
                'descricao_produto': produto.get('descricao_produto'),
                'has_image': 'caminho_imagem' in produto and produto['caminho_imagem'] is not None
            })
        
        return jsonify(results)

    except Exception as e:
        print(f"Erro na rota /search_products_for_upload: {str(e)}")
        return jsonify({'error': 'Erro ao buscar produtos'}), 500

@app.route('/upload_produtos', methods=['POST'])
@login_required
def upload_produtos():
    if current_user.role != 'admin':
        flash('Acesso não autorizado')
        return redirect(url_for('galeria'))
        
    if produtos_collection is None:
         flash('Erro: Conexão com o MongoDB não estabelecida.')
         return redirect(url_for('admin'))

    if 'file' not in request.files:
        flash('Nenhum arquivo selecionado')
        return redirect(url_for('admin'))
    
    file = request.files['file']
    if file.filename == '':
        flash('Nenhum arquivo selecionado')
        return redirect(url_for('admin'))
    
    if file and file.filename.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Ler cabeçalho para mapear colunas
            header = [cell.value for cell in ws[1]]
            
            produtos_para_importar = []
            
            for row in ws.iter_rows(min_row=2):  # Pula o cabeçalho
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(header):
                         # Limpa o nome da coluna para usar como chave no MongoDB
                        column_name = str(header[i]).replace('.', '_').replace(' ', '_') if header[i] else f'col_{i}'
                        row_data[column_name] = cell.value
                
                # Mapear colunas específicas e garantir campos essenciais
                codigo = row_data.get('Codigo', row_data.get('codigo_produto'))
                descricao = row_data.get('Descricao', row_data.get('descricao_produto'))
                # Tenta obter status de diferentes nomes de coluna comuns, default FALSE
                status_val = row_data.get('status')
                if status_val is None:
                     status_val = row_data.get('Ativo') # Tenta a coluna 'Ativo' também
                
                status_str = str(status_val).upper() if status_val is not None else 'FALSE'

                # Converter status para booleano
                ativo_catalogo = True if status_str == 'TRUE' else False
                
                if codigo and descricao:
                    # Preparar dados para upsert
                    update_data = {
                        'codigo_produto': codigo,
                        'descricao_produto': descricao,
                        'ativo_catalogo': ativo_catalogo,
                        # Adicionar todos os dados da linha lida, sobrescrevendo mapeamentos específicos
                        **row_data
                    }
                    
                    produtos_para_importar.append(update_data)
            
            # Realizar upsert em massa (opcional, mas mais eficiente para muitos produtos)
            # Para simplificar, faremos upsert um a um por enquanto.
            for produto_data in produtos_para_importar:
                 # Remove o campo _id se existir para evitar erro no upsert de novos documentos
                 produto_data.pop('_id', None)
                 
                 produtos_collection.update_one(
                    {'codigo_produto': produto_data['codigo_produto']},
                    {'$set': produto_data},
                    upsert=True
                )
        
            flash('Produtos importados com sucesso')
        except Exception as e:
            flash(f'Erro ao importar produtos: {str(e)}')
            print(f"Erro na rota /upload_produtos: {str(e)}")
    else:
        flash('Formato de arquivo inválido. Por favor, use um arquivo .xlsx')
    
    # Redirecionar para a primeira página do admin após importação
    return redirect(url_for('admin', page=1))

# Função para otimizar e salvar imagem
def otimizar_e_salvar_imagem(file_storage, caminho_destino, tamanho_max_kb=500):
    """Otimiza a imagem do upload, converte para JPEG, fundo branco se PNG, e salva."""
    img = Image.open(file_storage)
    # Se for PNG com transparência, converte para fundo branco
    if img.mode in ("RGBA", "P"):
        fundo = Image.new("RGB", img.size, (255, 255, 255))
        fundo.paste(img, mask=img.split()[-1])
        img = fundo
    else:
        img = img.convert("RGB")
    qualidade = 85
    while qualidade >= 30:
        img.save(caminho_destino, format='JPEG', quality=qualidade, optimize=True)
        tamanho_kb = os.path.getsize(caminho_destino) // 1024
        if tamanho_kb <= tamanho_max_kb:
            break
        qualidade -= 5

@app.route('/upload_imagem', methods=['POST'])
@login_required
def upload_imagem():
    if current_user.role != 'admin':
        flash('Acesso não autorizado')
        return redirect(url_for('galeria'))

    if produtos_collection is None:
         flash('Erro: Conexão com o MongoDB não estabelecida.')
         return redirect(url_for('admin'))
        
    # Alterado para pegar o codigo_produto do campo oculto após a busca
    codigo_produto = request.form.get('codigo_produto_upload')
    file = request.files.get('file')

    if not file or not codigo_produto:
        flash('Arquivo ou código do produto não especificado.')
        return redirect(url_for('admin'))
    
    if file and allowed_file(file.filename):
        try:
            # Verifica se o produto existe
            produto_existente = produtos_collection.find_one({'codigo_produto': codigo_produto})
            if not produto_existente:
                 flash(f'Código do produto {codigo_produto} não encontrado.')
                 return redirect(url_for('admin'))
                 
            # Sempre salva como JPEG otimizado
            filename = secure_filename(f"{codigo_produto}.jpg")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Garante que o diretório de uploads exista
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            
            # Otimiza e salva a imagem
            otimizar_e_salvar_imagem(file, filepath)
            
            # Caminho para salvar no banco (sempre .jpg agora)
            caminho_para_db = os.path.join('uploads', 'images', filename).replace('\\', '/')

            produtos_collection.update_one(
                {'codigo_produto': codigo_produto},
                {'$set': {'caminho_imagem': caminho_para_db}}
            )
            
            flash(f'Imagem enviada com sucesso para o produto {codigo_produto}.')
        except Exception as e:
            flash(f'Erro ao enviar imagem: {str(e)}')
            print(f"Erro na rota /upload_imagem: {str(e)}")
    else:
        flash('Formato de arquivo inválido')
    
    # Redirecionar de volta para a primeira página do admin após upload
    return redirect(url_for('admin', page=1))

@app.route('/salvar_decisoes', methods=['POST'])
@login_required
def salvar_decisoes():
    if current_user.role != 'dono':
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('galeria'))
    
    try:
        # Obter todos os campos do formulário
        form_data = request.form.to_dict()
        
        # Atualizar ativo_catalogo para todos os produtos exibidos
        todos_codigos = request.form.getlist('todos_codigos_produto[]')
        for codigo_produto in todos_codigos:
            ativo_catalogo = f'produto_{codigo_produto}' in form_data
            # Buscar o codigo_base dessa variante
            doc = produtos_collection.find_one({'codigo_produto': codigo_produto})
            if doc and 'codigo_base' in doc:
                codigo_base = doc['codigo_base']
                # Atualizar todas as variantes do grupo
                produtos_collection.update_many(
                    {'codigo_base': codigo_base},
                    {'$set': {'ativo_catalogo': ativo_catalogo}}
                )
            else:
                # Fallback: atualiza só o produto se não tiver codigo_base
                produtos_collection.update_one(
                    {'codigo_produto': codigo_produto},
                    {'$set': {'ativo_catalogo': ativo_catalogo}}
                )
        
        # Processar alterações de preço
        for key, value in form_data.items():
            if key.startswith('preco_'):
                # Formato esperado: preco_TABELA_CODIGO
                parts = key.split('_')
                if len(parts) >= 3:
                    tabela = '_'.join(parts[1:-1])  # Junta todas as partes do meio para formar o nome da tabela
                    codigo_produto = parts[-1]
                    try:
                        valor_str = str(value).replace('.', '').replace(',', '.')
                        valor = float(valor_str)
                        # Atualizar o preço na tabela específica
                        produtos_collection.update_one(
                            {'codigo_produto': codigo_produto},
                            {'$set': {tabela: valor}}
                        )
                    except ValueError:
                        continue  # Ignora valores inválidos
        
        flash('Decisões salvas com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar decisões: {str(e)}', 'danger')
    
    # Redirecionar de volta para a galeria mantendo os parâmetros da página
    current_page = request.form.get('current_page', 1)
    search_query = request.form.get('search_query', '')
    tabela_preco = request.form.get('tabela_preco', '')
    return redirect(url_for('galeria', page=current_page, search=search_query, tabela_preco=tabela_preco))

@app.route('/exportar_protheus')
@login_required
def exportar_protheus():
    if current_user.role != 'admin':
        flash('Acesso não autorizado')
        return redirect(url_for('galeria'))

    if produtos_collection is None:
         flash('Erro: Conexão com o MongoDB não estabelecida.')
         return redirect(url_for('admin'))
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Cabeçalho - Incluindo todas as colunas do DB
        # Vamos tentar obter o cabeçalho de um produto existente para incluir todas as chaves
        sample_produto = produtos_collection.find_one()
        if sample_produto:
            # Excluir _id do cabeçalho
            header = [key for key in sample_produto.keys() if key != '_id']
            # Garantir que as colunas principais estejam no início
            main_columns = ['codigo_produto', 'descricao_produto', 'ativo_catalogo', 'Preco_CSL']
            sorted_header = main_columns + [col for col in header if col not in main_columns]
            ws.append(sorted_header)

            # Dados
            for produto in produtos_collection.find():
                row_data = []
                for col in sorted_header:
                     # Formatar 'ativo_catalogo' como 'ATIVO'/'INATIVO' para exportação
                    if col == 'ativo_catalogo':
                        row_data.append('ATIVO' if produto.get(col, False) else 'INATIVO')
                    else:
                         row_data.append(produto.get(col, '')) # Use get para evitar KeyError
                ws.append(row_data)
        else: # Caso não haja produtos, apenas cria um cabeçalho básico
             ws.append(['Código do Produto', 'Descrição do Produto', 'Status Catálogo', 'Preco CSL'])
        
        # Salvar arquivo
        filename = f'catalogo_protheus_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        # Cria o diretório 'uploads' se não existir antes de salvar
        os.makedirs('uploads', exist_ok=True)
        filepath = os.path.join('uploads', filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        flash(f'Erro ao exportar dados: {str(e)}')
        print(f"Erro na rota /exportar_protheus: {str(e)}")
        return redirect(url_for('admin'))

@app.route('/produto/<codigo_base>')
def produto_detalhe(codigo_base):
    if produtos_collection is None:
        flash('Erro: Conexão com o MongoDB não estabelecida.')
        return redirect(url_for('galeria'))

    # Buscar todas as variantes desse codigo_base
    variantes = list(produtos_collection.find({
        '$or': [
            {'codigo_base': codigo_base},
            # fallback para produtos antigos sem campo codigo_base
            {'codigo_produto': {'$regex': f'^{codigo_base}'}},
        ]
    }))
    if not variantes:
        flash('Produto não encontrado.')
        return redirect(url_for('galeria'))

    # Pega o produto principal (primeira variante)
    produto = variantes[0]

    # Listar todas as cores disponíveis
    cores_disponiveis = [v.get('Desc_Cor', 'N/A') for v in variantes]

    # Listar todos os preços de todas as tabelas para cada variante
    tabelas_preco = [
        'TABELA_VAREJO_5_000',
        'TABELA_VAREJO_25_000',
        'TABELA_VAREJO_50_000',
        'TABELA_ATACADO',
        'TABELA_GOLD_PARCEIRO',
        'TABELA_ATACADO_AMAZONIA_E_MACAPA',
        'TABELA_ACRE_E_RONDONIA'
    ]

    return render_template(
        'produto.html',
        produto=produto,
        variantes=variantes,
        cores_disponiveis=cores_disponiveis,
        tabelas_preco=tabelas_preco
    )

@app.template_filter('format_price')
def format_price(value):
    try:
        return '{:,.2f}'.format(float(str(value).replace(',', '.'))).replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return '0,00'

@app.route('/editar_precos/<codigo_base>', methods=['GET', 'POST'])
@login_required
def editar_precos(codigo_base):
    if current_user.role != 'dono':
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('galeria'))

    if request.method == 'POST':
        # Salvar os preços editados
        tabelas_preco = [
            'TABELA_VAREJO_5_000',
            'TABELA_VAREJO_25_000',
            'TABELA_VAREJO_50_000',
            'TABELA_ATACADO',
            'TABELA_GOLD_PARCEIRO',
            'TABELA_ATACADO_AMAZONIA_E_MACAPA',
            'TABELA_ACRE_E_RONDONIA'
        ]
        variantes = list(produtos_collection.find({'codigo_base': codigo_base}))
        for variante in variantes:
            codigo_produto = variante['codigo_produto']
            updates = {}
            for tabela in tabelas_preco:
                key = f'{tabela}_{codigo_produto}'
                valor = request.form.get(key)
                if valor is not None:
                    try:
                        valor_str = str(valor).replace('.', '').replace(',', '.')
                        updates[tabela] = float(valor_str)
                    except Exception:
                        continue
            if updates:
                produtos_collection.update_one({'codigo_produto': codigo_produto}, {'$set': updates})
        flash('Preços atualizados com sucesso!', 'success')
        return redirect(url_for('galeria'))

    # GET: mostrar a tabela
    variantes = list(produtos_collection.find({'codigo_base': codigo_base}))
    descricao_produto = variantes[0]['descricao_produto'] if variantes else ''
    tabelas_preco = [
        'TABELA_VAREJO_5_000',
        'TABELA_VAREJO_25_000',
        'TABELA_VAREJO_50_000',
        'TABELA_ATACADO',
        'TABELA_GOLD_PARCEIRO',
        'TABELA_ATACADO_AMAZONIA_E_MACAPA',
        'TABELA_ACRE_E_RONDONIA'
    ]
    return render_template('editar_precos.html', variantes=variantes, tabelas_preco=tabelas_preco, codigo_base=codigo_base, descricao_produto=descricao_produto)

@app.route('/exportar_catalogo_pdf', methods=['GET', 'POST'])
@login_required
def exportar_catalogo_pdf():
    if current_user.role not in ['admin', 'dono']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('galeria'))

    tabelas_preco = [
        ('TABELA_VAREJO_5_000', 'Varejo 5k'),
        ('TABELA_VAREJO_25_000', 'Varejo 25k'),
        ('TABELA_VAREJO_50_000', 'Varejo 50k'),
        ('TABELA_ATACADO', 'Atacado'),
        ('TABELA_GOLD_PARCEIRO', 'Gold Parceiro'),
        ('TABELA_ATACADO_AMAZONIA_E_MACAPA', 'Atacado Amaz/Mac'),
        ('TABELA_ACRE_E_RONDONIA', 'Acre/Rondônia')
    ]

    if request.method == 'POST':
        try:
            tabela_preco = request.form.get('tabela_preco')
            apenas_com_foto = request.form.get('apenas_com_foto') == 'on'

            # 1. Buscar TODAS as variantes que são 'ativo_catalogo: true'
            filtro_inicial = {'ativo_catalogo': True}
            todas_variantes_ativas = list(produtos_collection.find(filtro_inicial))

            # 2. Agrupar estas variantes por codigo_base
            produtos_por_grupo_temp = {}
            for variante_ativa in todas_variantes_ativas:
                codigo_base = variante_ativa.get('codigo_base')
                if not codigo_base:  # Pular variantes sem codigo_base
                    continue
                if codigo_base not in produtos_por_grupo_temp:
                    produtos_por_grupo_temp[codigo_base] = []
                produtos_por_grupo_temp[codigo_base].append(variante_ativa)

            # 3. Preparar uma lista de grupos
            grupos_pre_filtro = []
            for codigo_base_key, variantes_do_grupo_atual in produtos_por_grupo_temp.items():
                if not variantes_do_grupo_atual:
                    continue

                # Determinar o produto principal (preferencialmente branco)
                produto_principal_do_grupo = None
                for v_principal_check in variantes_do_grupo_atual:
                    if v_principal_check.get('codigo_produto','').endswith('1'):
                        produto_principal_do_grupo = v_principal_check
                        break
                if not produto_principal_do_grupo:
                    produto_principal_do_grupo = sorted(variantes_do_grupo_atual, key=lambda v_sort: v_sort.get('codigo_produto', ''))[0]
                
                variantes_ordenadas_para_template = sorted(variantes_do_grupo_atual, key=lambda v_sort: v_sort.get('codigo_produto', ''))

                # Limpar a descrição
                desc_original = produto_principal_do_grupo.get('descricao_produto', '')
                desc_temp = desc_original
                palavras_chave_remover = ["branco", "branca", "branc"]

                for palavra in palavras_chave_remover:
                    regex_pattern_direita = r'\b' + re.escape(palavra) + r'\b\s+'
                    desc_temp = re.sub(regex_pattern_direita, '', desc_temp, flags=re.IGNORECASE)
                    regex_pattern_esquerda = r'\s+\b' + re.escape(palavra) + r'\b'
                    desc_temp = re.sub(regex_pattern_esquerda, '', desc_temp, flags=re.IGNORECASE)
                    if desc_temp.lower() == palavra.lower():
                        desc_temp = ''
                
                desc_final = ' '.join(desc_temp.split()).strip()

                grupos_pre_filtro.append({
                    'descricao_produto': desc_final,
                    'caminho_imagem': produto_principal_do_grupo.get('caminho_imagem'),
                    'codigo_base': codigo_base_key,
                    'variants': variantes_ordenadas_para_template,
                    '_tem_imagem_principal': True if produto_principal_do_grupo.get('caminho_imagem') else False
                })

            # 4. Filtrar os GRUPOS se 'apenas_com_foto' estiver marcado
            produtos_agrupados = []
            if apenas_com_foto:
                for grupo_candidato in grupos_pre_filtro:
                    if grupo_candidato['_tem_imagem_principal']:
                        del grupo_candidato['_tem_imagem_principal']
                        produtos_agrupados.append(grupo_candidato)
            else:
                for grupo_candidato in grupos_pre_filtro:
                    del grupo_candidato['_tem_imagem_principal']
                    produtos_agrupados.append(grupo_candidato)
            
            produtos_agrupados.sort(key=lambda g: g.get('descricao_produto',''))

            # Adicionar flag para tratamento especial
            for grupo in produtos_agrupados:
                grupo['is_special_bpb_product'] = any(v.get('tratamento_cor_especial', False) for v in grupo['variants'])

            if HTML is None:
                flash('WeasyPrint não está instalado no servidor.', 'danger')
                return redirect(url_for('exportar_catalogo_pdf'))

            # Gerar o PDF
            html = render_template('catalogo_pdf.html', 
                                 produtos=produtos_agrupados, 
                                 tabela_preco=tabela_preco, 
                                 tabelas_preco=tabelas_preco)

            # Criar um nome de arquivo único
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'catalogo_luzarte_{timestamp}.pdf'
            
            # Criar diretório temporário se não existir
            temp_dir = os.path.join('static', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Caminho completo do arquivo
            filepath = os.path.join(temp_dir, filename)
            
            # Gerar o PDF
            HTML(string=html, base_url=request.base_url).write_pdf(filepath)
            
            # Enviar o arquivo
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/pdf'
            )

        except Exception as e:
            flash(f'Erro ao gerar o PDF: {str(e)}', 'danger')
            print(f"Erro na geração do PDF: {str(e)}")
            return redirect(url_for('exportar_catalogo_pdf'))

    return render_template('exportar_catalogo_pdf.html', tabelas_preco=tabelas_preco)

@app.route('/api/search_suggestions')
@login_required
def search_suggestions():
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify([])
    
    try:
        # Buscar produtos que correspondam à query
        produtos = list(produtos_collection.find({
            'descricao_produto': {'$regex': re.escape(query), '$options': 'i'}
        }).limit(10))
        
        # Formatar resultados
        suggestions = []
        for produto in produtos:
            suggestions.append({
                'id': produto['codigo_produto'],
                'text': f"{produto['descricao_produto']} ({produto['codigo_produto']})",
                'descricao': produto['descricao_produto'],
                'codigo': produto['codigo_produto']
            })
        
        return jsonify(suggestions)
    except Exception as e:
        print(f"Erro ao buscar sugestões: {str(e)}")
        return jsonify([])

@app.route('/catalogo')
@login_required
def catalogo():
    # Mesma lógica da galeria, mas sempre modo visualização (sem edição)
    if produtos_collection is None:
         flash('Erro: Conexão com o MongoDB não estabelecida.')
         return render_template('catalogo.html', produtos=[], total_pages=0, current_page=1, search_query='', tabela_preco_selecionada='TABELA_VAREJO_5_000', categorias=[], cores=[])

    page = request.args.get('page', 1, type=int)
    search_query = request.args.get('search', '')
    tabela_preco_selecionada = request.args.get('tabela_preco', 'TABELA_VAREJO_5_000')
    categoria_selecionada = request.args.get('categoria', '')
    cor_selecionada = request.args.get('cor', '')
    per_page = 24

    try:
        categorias = produtos_collection.distinct('Desc_Familia')
        categorias = [cat for cat in categorias if cat]
        categorias.sort()
    except Exception as e:
        print(f"Erro ao buscar categorias: {str(e)}")
        categorias = []
    try:
        cores = produtos_collection.distinct('Desc_Cor')
        cores = [cor for cor in cores if cor]
        cores.sort()
    except Exception as e:
        print(f"Erro ao buscar cores: {str(e)}")
        cores = []

    pipeline = []
    match_query = {'ativo_catalogo': True}
    if search_query:
        match_query['descricao_produto'] = {'$regex': re.escape(search_query), '$options': 'i'}
    if categoria_selecionada:
        match_query['Desc_Familia'] = categoria_selecionada
    pipeline.append({'$match': match_query})
    pipeline.append({'$addFields': {
        'codigo_base': { '$substrCP': ['$codigo_produto', 0, { '$subtract': [{'$strLenCP': '$codigo_produto'}, 1] }] }
    }})
    pipeline.append({
        '$group': {
            '_id': '$codigo_base',
            'variants': { '$push': '$$ROOT' },
            'main_product': {
                '$first': {
                    '$filter': {
                        'input': '$variants',
                        'as': 'variant',
                        'cond': { '$or': [
                            { '$eq': [{ '$substrCP': [{'$ifNull': ['$$variant.codigo_produto', '']}, { '$subtract': [{'$strLenCP': {'$ifNull': ['$$variant.codigo_produto', '']}}, 1] }, 1] }, '1'] },
                            { '$eq': [{'$ifNull': ['$$variant.Desc_Cor', '']}, 'BRANCO'] }
                        ]}
                    }
                }
            }
        }
    })
    pipeline.append({
        '$project': {
            '_id': 0,
            'codigo_base': '$_id',
            'main_product': { '$ifNull': ['$main_product', { '$arrayElemAt': ['$variants', 0] }] },
            'variants': '$variants'
        }
    })
    count_pipeline = pipeline[:-1]
    count_pipeline.append({'$count': 'total'})
    total_produtos_agrupados = 0
    try:
        count_result = list(produtos_collection.aggregate(count_pipeline))
        if count_result:
            total_produtos_agrupados = count_result[0].get('total', 0)
    except Exception as e:
        print(f"Erro ao contar produtos agrupados: {str(e)}")
    total_pages = math.ceil(total_produtos_agrupados / per_page)
    if page < 1:
        page = 1
    elif page > total_pages and total_pages > 0:
        page = total_pages
    elif total_pages == 0:
         page = 1
    skip = (page - 1) * per_page
    pipeline.append({'$skip': skip})
    pipeline.append({'$limit': per_page})
    produtos_agrupados = []
    try:
        produtos_agrupados = list(produtos_collection.aggregate(pipeline))
        produtos_para_template = []
        for grupo in produtos_agrupados:
            if cor_selecionada:
                if not any(variant.get('Desc_Cor') == cor_selecionada for variant in grupo['variants']):
                    continue
            main_prod = grupo['main_product']
            variants_list = []
            for variant in grupo['variants']:
                tabelas_preco = [
                    'TABELA_VAREJO_5_000',
                    'TABELA_VAREJO_25_000',
                    'TABELA_VAREJO_50_000',
                    'TABELA_ATACADO',
                    'TABELA_GOLD_PARCEIRO',
                    'TABELA_ATACADO_AMAZONIA_E_MACAPA',
                    'TABELA_ACRE_E_RONDONIA'
                ]
                precos = {k: variant.get(k, 0) for k in tabelas_preco}
                variant_data = {
                    'codigo_produto': variant.get('codigo_produto'),
                    'Desc_Cor': variant.get('Desc_Cor', 'Não Informado'),
                    'precos': precos
                }
                variants_list.append(variant_data)
            produtos_para_template.append({
                'codigo_base': grupo['codigo_base'],
                'descricao_produto': main_prod.get('descricao_produto'),
                'caminho_imagem': main_prod.get('caminho_imagem'),
                'ativo_catalogo': main_prod.get('ativo_catalogo', False),
                'variants': variants_list
            })
        return render_template('catalogo.html',
                               produtos=produtos_para_template,
                               total_pages=total_pages,
                               current_page=page,
                               search_query=search_query,
                               tabela_preco_selecionada=tabela_preco_selecionada,
                               categorias=categorias,
                               categoria_selecionada=categoria_selecionada,
                               cores=cores,
                               cor_selecionada=cor_selecionada)
    except Exception as e:
        print("ERRO GRAVE NO CATALOGO:", e)
        import traceback
        traceback.print_exc()
        flash(f'Erro ao carregar produtos no catálogo: {str(e)}')
        return render_template('catalogo.html', produtos=[], total_pages=0, current_page=1, search_query=search_query, tabela_preco_selecionada=tabela_preco_selecionada, categorias=categorias, categoria_selecionada=categoria_selecionada, cores=cores, cor_selecionada=cor_selecionada)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True) 